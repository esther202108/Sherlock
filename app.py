import streamlit as st
import pandas as pd
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────────────────────────────────────────────────────
# Streamlit setup (more professional UX)
# ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="Sherlock - Compare Vendor Files", layout="wide")
st.title("🕵️‍♂️ Sherlock")
st.caption(
    "Spotting what’s changed, one name at a time.\n"
    "Same vendor. Different clues."
)

NAME_COL = "Full Name As Per NRIC"
SERIAL_COL = "S/N"

# Fixed column widths (EXACT match to US Cleaner)
COLUMN_WIDTHS = {
    "A": 3.38,   # S/N
    "C": 23.06,
    "D": 25,
    "E": 17.63,
    "F": 26.25,
    "G": 13.94,
    "H": 24.06,
    "I": 18.38,
    "J": 20.31,
    "K": 4,
    "L": 5.81,
    "M": 11.5,
}

# Common serial column variants (for input files)
SERIAL_CANDIDATES = [
    "S/N", "SN", "SNO", "S. NO", "S. NO.", "S NO", "S NO.", "NO", "NO.",
    "INDEX", "SERIAL", "SERIAL NO", "SERIAL NO."
]

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
def pick_sheet(file, key_prefix: str):
    ext = file.name.lower().split(".")[-1]
    engine = "openpyxl" if ext == "xlsx" else "xlrd"
    xl = pd.ExcelFile(file, engine=engine)
    sheet = st.selectbox(
        f"Sheet ({file.name})",
        xl.sheet_names,
        key=f"{key_prefix}_sheet"
    )
    return xl.parse(sheet), sheet

def normalize_name(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .fillna("")
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
        .str.upper()
    )

def detect_serial_col(df: pd.DataFrame) -> str | None:
    # Try exact match first
    for c in df.columns:
        if str(c).strip().upper() in SERIAL_CANDIDATES:
            return c
    return None

def filter_real_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Removes footer/summary rows like 'Vehicles' / 'Total Visitors' by keeping only rows
    where the detected serial column is numeric.
    If no serial column exists, keeps df as-is.
    """
    serial_col = detect_serial_col(df)
    if not serial_col:
        return df.copy()

    s = pd.to_numeric(df[serial_col], errors="coerce")
    mask = s.notna()  # only numeric rows
    return df.loc[mask].copy()

def count_real_records(df: pd.DataFrame) -> int:
    """
    Counts only real rows:
    - If serial column exists: count numeric serial cells
    - Else: count non-empty names (fallback)
    """
    serial_col = detect_serial_col(df)
    if serial_col:
        s = pd.to_numeric(df[serial_col], errors="coerce")
        return int(s.notna().sum())

    # Fallback: count non-empty names
    if NAME_COL in df.columns:
        n = normalize_name(df[NAME_COL])
        return int((n != "").sum())

    return int(len(df))

def add_serial_number(df: pd.DataFrame) -> pd.DataFrame:
    df = df.reset_index(drop=True).copy()

    # Remove any existing serial column (common variations)
    for c in list(df.columns):
        if str(c).strip().upper() in SERIAL_CANDIDATES:
            df.drop(columns=[c], inplace=True)

    df.insert(0, SERIAL_COL, range(1, len(df) + 1))
    return df

# ─────────────────────────────────────────────────────────────
# Excel styling (MATCH US CLEANER)
# ─────────────────────────────────────────────────────────────
def apply_us_style(ws):
    header_fill = PatternFill("solid", fgColor="94B455")
    border = Border(Side("thin"), Side("thin"), Side("thin"), Side("thin"))
    center = Alignment(horizontal="center", vertical="center")
    normal_font = Font(name="Calibri", size=9)
    bold_font = Font(name="Calibri", size=9, bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center
            cell.font = normal_font

    for col in range(1, ws.max_column + 1):
        cell = ws[f"{get_column_letter(col)}1"]
        cell.fill = header_fill
        cell.font = bold_font

    ws.freeze_panes = "A2"

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

def apply_fixed_column_widths(ws):
    ws.column_dimensions["A"].width = COLUMN_WIDTHS.get("A", 3.38)
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def build_workbook(sheets: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        apply_us_style(ws)
        apply_fixed_column_widths(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────
# UX Layout
# ─────────────────────────────────────────────────────────────
with st.expander("How it works", expanded=False):
    st.markdown(
        """
        - **Excel A (Baseline):** Previous roster  
        - **Excel B (Current):** Current roster
        - Output includes:
          - **New Personnel(s) in Excel B**
          - **Removed Personnel(s) from Excel A**
        """
    )

with st.container(border=True):
    st.subheader("1) Upload files")
    c1, c2 = st.columns(2)
    with c1:
        file_a = st.file_uploader("Excel A (Baseline)", type=["xlsx", "xls"])
    with c2:
        file_b = st.file_uploader("Excel B (Compare)", type=["xlsx", "xls"])

if file_a and file_b:
    with st.container(border=True):
        st.subheader("2) Select sheets")
        c1, c2 = st.columns(2)
        with c1:
            df_a_raw, sheet_a = pick_sheet(file_a, "a")
        with c2:
            df_b_raw, sheet_b = pick_sheet(file_b, "b")

        # ✅ filter out footer rows (Vehicles / Total Visitors etc.)
        df_a = filter_real_rows(df_a_raw)
        df_b = filter_real_rows(df_b_raw)

        # Validation
        issues = []
        if NAME_COL not in df_a.columns:
            issues.append(f"Excel A is missing column: **{NAME_COL}**")
        if NAME_COL not in df_b.columns:
            issues.append(f"Excel B is missing column: **{NAME_COL}**")

        if issues:
            for msg in issues:
                st.error(msg)
            st.stop()

    # Compare (only real rows)
    a_norm = normalize_name(df_a[NAME_COL])
    b_norm = normalize_name(df_b[NAME_COL])

    a_set = set(a_norm[a_norm != ""].tolist())
    b_set = set(b_norm[b_norm != ""].tolist())

    new_set = b_set - a_set
    removed_set = a_set - b_set

    new_rows = df_b.loc[b_norm.isin(new_set)].copy()
    removed_rows = df_a.loc[a_norm.isin(removed_set)].copy()

    new_out = add_serial_number(new_rows)
    removed_out = add_serial_number(removed_rows)

    # KPIs (✅ count only real roster rows / numeric S/N)
    st.subheader("3) Summary")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Baseline Counts (A)", f"{count_real_records(df_a_raw):,}")
    k2.metric("Current Counts (B)", f"{count_real_records(df_b_raw):,}")
    k3.metric("New Personnel(s) in B", f"{len(new_out):,}")
    k4.metric("Removed Personnel(s) from A", f"{len(removed_out):,}")

    # Results Preview
    with st.container(border=True):
        st.subheader("4) Results (Preview)")
        col_cfg = {
            SERIAL_COL: st.column_config.NumberColumn(SERIAL_COL, width="xsmall"),
            NAME_COL: st.column_config.TextColumn(NAME_COL, width="large"),
        }

        l, r = st.columns(2, gap="large")
        with l:
            st.markdown("#### 🆕 New Personnel(s) in Excel B")
            if not new_out.empty:
                st.dataframe(
                    new_out[[SERIAL_COL, NAME_COL]],
                    hide_index=True,
                    use_container_width=True,
                    column_config=col_cfg,
                )
            else:
                st.success("No new names found in Excel B.")

        with r:
            st.markdown("#### ❌ Removed Personnel(s) from Excel A")
            if not removed_out.empty:
                st.dataframe(
                    removed_out[[SERIAL_COL, NAME_COL]],
                    hide_index=True,
                    use_container_width=True,
                    column_config=col_cfg,
                )
            else:
                st.success("No names removed from Excel A.")

    # Download
    with st.container(border=True):
        st.subheader("5) Export")
        st.caption("Downloads an Excel file")

        output = build_workbook({
            "New_in_Excel_B": new_out,
            "Removed_from_Excel_A": removed_out,
        })

        st.download_button(
            "📥 Download Results (Styled XLSX)",
            data=output,
            file_name="clearid_compare.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

else:
    st.info("Upload both files to begin.")
